"""
make_xlsx.py — Convert validated submission.csv to submission.xlsx.

Reads submission.csv exactly as-is, writes an identical XLSX,
then verifies row-by-row that every field matches.

Usage:
    python make_xlsx.py [--csv submission.csv] [--xlsx submission.xlsx]
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Write an XLSX file from a CSV, preserving all values exactly."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission"

    # Header styling
    header_font = Font(bold=True, name="Calibri", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Column widths (characters)
    col_widths = {
        "candidate_id": 18,
        "rank":         8,
        "score":        18,
        "reasoning":    90,
    }

    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                else:
                    # Keep score as string to preserve exact decimal representation
                    cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 4))

    # Apply column widths
    for col_idx, header in enumerate(col_widths.keys(), start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths[header]

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"[make_xlsx] Written: {xlsx_path} ({ws.max_row - 1} data rows)")


def verify_identical(csv_path: Path, xlsx_path: Path) -> bool:
    """
    Row-by-row comparison of CSV and XLSX.
    Returns True if identical, False + prints diffs if not.
    """
    # Read CSV
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.reader(f))

    # Read XLSX
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    xlsx_rows = []
    for row in ws.iter_rows(values_only=True):
        xlsx_rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    if len(csv_rows) != len(xlsx_rows):
        print(f"[FAIL] Row count mismatch: CSV={len(csv_rows)}, XLSX={len(xlsx_rows)}")
        return False

    diffs = 0
    for i, (cr, xr) in enumerate(zip(csv_rows, xlsx_rows), start=1):
        if cr != xr:
            diffs += 1
            print(f"[DIFF] Row {i}:")
            print(f"  CSV : {cr}")
            print(f"  XLSX: {xr}")

    if diffs == 0:
        print(f"[make_xlsx] Verification PASSED: {len(csv_rows)} rows, zero differences.")
        return True
    else:
        print(f"[make_xlsx] Verification FAILED: {diffs} differing rows.")
        return False


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert submission.csv → submission.xlsx")
    parser.add_argument("--csv",  default="submission.csv",  help="Input CSV path")
    parser.add_argument("--xlsx", default="submission.xlsx", help="Output XLSX path")
    args = parser.parse_args()

    csv_path  = Path(args.csv)
    xlsx_path = Path(args.xlsx)

    if not csv_path.exists():
        print(f"Error: {csv_path} not found", file=sys.stderr)
        sys.exit(1)

    csv_to_xlsx(csv_path, xlsx_path)
    ok = verify_identical(csv_path, xlsx_path)
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
